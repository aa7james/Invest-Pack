"""
Incremental sync: read the Excel, compare with what's already in Supabase, and
upsert ONLY the datapoints that are new or changed. Much faster than a full
re-load (which rewrites all ~480k rows).

Run via "SYNC DATA.bat" (prompts for the secret key), or:
    py watcher/sync_history.py
"""
import os
import sys
import json
import datetime
import urllib.request
import urllib.error

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl missing. Run:  py -m pip install openpyxl")
    sys.exit(1)

HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_EXCEL = r"C:\Users\JamesWhitaker\Aylett & Co\Research - Documents\Process\New order\4. Drivers\Data Dump - HP.xlsx"
CAT_MAP = {
    "1.BB Data - Chemicals": "Chemicals", "2.BB Data - Energy": "Energy",
    "3.BB Data - Metals": "Metals", "4.BB Data - Soft Commodities": "Soft Commodities",
    "5. BB Data - Chicken Dashboard": "Chicken", "7.BB Data - Gaming": "Gaming",
    "9.BB Data - Manheim": "Manheim",
}
BATCH = 2000
EPS = 0.005  # values within this are treated as unchanged


def load_env():
    env = {}
    path = os.path.join(HERE, ".env")
    if os.path.exists(path):
        for line in open(path, "r", encoding="utf-8"):
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    url = (os.environ.get("SUPABASE_URL") or env.get("SUPABASE_URL", "")).rstrip("/")
    key = os.environ.get("SUPABASE_SECRET_KEY") or env.get("SUPABASE_SECRET_KEY", "")
    excel = os.environ.get("EXCEL_PATH") or env.get("EXCEL_PATH", DEFAULT_EXCEL)
    if not url or not key or "paste" in key.lower():
        print("ERROR: SUPABASE_URL / SUPABASE_SECRET_KEY not provided.")
        sys.exit(1)
    return url, key, excel


def api(method, url, key, path, body=None, extra=None):
    headers = {"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json",
               "Accept-Profile": "pack", "Content-Profile": "pack"}
    if extra:
        headers.update(extra)
    data = json.dumps(body).encode("utf-8") if body is not None else None
    req = urllib.request.Request(url + path, data=data, method=method, headers=headers)
    with urllib.request.urlopen(req, timeout=180) as resp:
        raw = resp.read().decode("utf-8")
        return json.loads(raw) if raw else None


def norm(t):
    return " ".join(str(t).split()) if t not in (None, "") else ""


def main():
    url, key, excel = load_env()
    if not os.path.exists(excel):
        print("ERROR: Excel not found:", excel); sys.exit(1)

    print("Mapping instruments...")
    insts = api("GET", url, key, "/rest/v1/instruments?select=id,category,name,bloomberg_ticker,currency&limit=10000")
    lookup = {}
    for r in insts:
        lookup[(r["category"], r["name"], norm(r.get("bloomberg_ticker")), r.get("currency") or "")] = r["id"]

    print("Melting Excel...")
    wb = openpyxl.load_workbook(excel, read_only=True, data_only=True)
    excel_pts = {}
    for sheet, cat in CAT_MAP.items():
        if sheet not in wb.sheetnames:
            continue
        data = list(wb[sheet].iter_rows(min_row=1, values_only=True))
        names, tickers, fxs = data[2], data[3], data[4]
        cols = [c for c in range(1, len(names)) if names[c] not in (None, "")]
        for row in data[7:]:
            d = row[0]
            if not isinstance(d, (datetime.datetime, datetime.date)):
                continue
            iso = (d.date() if isinstance(d, datetime.datetime) else d).isoformat()
            for c in cols:
                v = row[c] if c < len(row) else None
                if not isinstance(v, (int, float)):
                    continue
                fx = (str(fxs[c]).replace("FX=", "").strip() if c < len(fxs) and fxs[c] else "")
                iid = lookup.get((cat, str(names[c]).strip(), norm(tickers[c] if c < len(tickers) else None), fx))
                if iid is not None:
                    excel_pts[(iid, iso)] = float(v)
    print(f"  {len(excel_pts):,} datapoints in Excel")

    print("Fetching current data from Supabase...")
    db = {}
    offset = 0
    while True:
        d = api("GET", url, key,
                f"/rest/v1/pack_data?select=instrument_id,obs_date,value&order=instrument_id,obs_date&limit=1000&offset={offset}")
        if not d:
            break
        for r in d:
            db[(r["instrument_id"], r["obs_date"])] = r["value"]
        if len(d) < 1000:
            break
        offset += 1000
    print(f"  {len(db):,} datapoints in database")

    changed = [{"instrument_id": iid, "obs_date": iso, "value": v}
               for (iid, iso), v in excel_pts.items()
               if (iid, iso) not in db or abs(db[(iid, iso)] - v) > EPS]

    if not changed:
        print("\nUp to date — nothing to upload.")
        return
    print(f"\n{len(changed):,} new/changed datapoints to upload.")
    total = 0
    for i in range(0, len(changed), BATCH):
        api("POST", url, key, "/rest/v1/pack_data?on_conflict=instrument_id,obs_date",
            body=changed[i:i + BATCH], extra={"Prefer": "resolution=merge-duplicates,return=minimal"})
        total += len(changed[i:i + BATCH])
        print(f"  ...uploaded {total:,}/{len(changed):,}")
    print(f"\nDONE. Synced {total:,} datapoints (left {len(db) - 0:,} untouched).")


if __name__ == "__main__":
    main()
