"""
One-time historical loader: reads the Bloomberg "Data Dump" Excel and loads all
history into pack.pack_data in Supabase.

Run it by double-clicking "LOAD HISTORY.bat" in the Invest-Pack folder, or:
    py watcher/load_history.py

It reads SUPABASE_URL + SUPABASE_SECRET_KEY from watcher/.env (the secret key
bypasses RLS so it can write price data). Safe to re-run — it upserts on
(instrument_id, obs_date), so nothing duplicates.
"""

import os
import sys
import json
import datetime
import urllib.request
import urllib.error

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is not installed. Run:  py -m pip install openpyxl")
    sys.exit(1)

HERE = os.path.dirname(os.path.abspath(__file__))

DEFAULT_EXCEL = r"C:\Users\JamesWhitaker\Aylett & Co\Research - Documents\Process\New order\4. Drivers\Data Dump - HP.xlsx"

# Sheet name -> category label (must match the seed).
CAT_MAP = {
    "1.BB Data - Chemicals": "Chemicals",
    "2.BB Data - Energy": "Energy",
    "3.BB Data - Metals": "Metals",
    "4.BB Data - Soft Commodities": "Soft Commodities",
    "5. BB Data - Chicken Dashboard": "Chicken",
    "7.BB Data - Gaming": "Gaming",
    "9.BB Data - Manheim": "Manheim",
}

BATCH = 5000


def load_env():
    # Prefer environment variables (set by the launcher's prompt); fall back to
    # watcher/.env so the future unattended watcher can use a saved key.
    env = {}
    path = os.path.join(HERE, ".env")
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()

    url = (os.environ.get("SUPABASE_URL") or env.get("SUPABASE_URL", "")).rstrip("/")
    key = os.environ.get("SUPABASE_SECRET_KEY") or env.get("SUPABASE_SECRET_KEY", "")
    excel = os.environ.get("EXCEL_PATH") or env.get("EXCEL_PATH", DEFAULT_EXCEL)

    if not url:
        print("ERROR: SUPABASE_URL is not set.")
        sys.exit(1)
    if not key or "paste" in key.lower():
        print("ERROR: No Supabase secret key provided.")
        sys.exit(1)
    return url, key, excel


def api(method, url, key, path, body=None, extra_headers=None):
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
        "Accept-Profile": "pack",   # read from the pack schema
        "Content-Profile": "pack",  # write to the pack schema
    }
    if extra_headers:
        headers.update(extra_headers)
    data = json.dumps(body).encode("utf-8") if body is not None else None
    req = urllib.request.Request(url + path, data=data, method=method, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            raw = resp.read().decode("utf-8")
            return json.loads(raw) if raw else None
    except urllib.error.HTTPError as e:
        print(f"HTTP {e.code} on {method} {path}: {e.read().decode('utf-8')[:500]}")
        raise


def norm_ticker(t):
    if t in (None, ""):
        return ""
    return " ".join(str(t).split())


def main():
    url, key, excel_path = load_env()
    if not os.path.exists(excel_path):
        print(f"ERROR: Excel file not found at:\n  {excel_path}")
        sys.exit(1)

    print("Fetching instruments from Supabase...")
    instruments = api("GET", url, key, "/rest/v1/instruments?select=id,category,name,bloomberg_ticker,currency&limit=10000")
    lookup = {}
    for r in instruments:
        lookup[(r["category"], r["name"], norm_ticker(r.get("bloomberg_ticker")), r.get("currency") or "")] = r["id"]
    print(f"  {len(lookup)} instruments mapped.")

    print(f"Reading Excel:\n  {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    rows_buffer = []
    total = 0
    unmatched = set()
    date_min, date_max = None, None

    def flush():
        nonlocal rows_buffer, total
        if not rows_buffer:
            return
        api("POST", url, key,
            "/rest/v1/pack_data?on_conflict=instrument_id,obs_date",
            body=rows_buffer,
            extra_headers={"Prefer": "resolution=merge-duplicates,return=minimal"})
        total += len(rows_buffer)
        print(f"  ...loaded {total:,} datapoints")
        rows_buffer = []

    for sheet, cat in CAT_MAP.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        data = list(ws.iter_rows(min_row=1, values_only=True))
        names = data[2]      # row 3 = display names
        tickers = data[3]    # row 4 = tickers
        fxs = data[4]        # row 5 = FX
        series_cols = [c for c in range(1, len(names)) if names[c] not in (None, "")]

        for r in data[7:]:   # data starts row 8
            d = r[0]
            if not isinstance(d, (datetime.datetime, datetime.date)):
                continue
            obs_date = d.date() if isinstance(d, datetime.datetime) else d
            iso = obs_date.isoformat()
            if date_min is None or iso < date_min:
                date_min = iso
            if date_max is None or iso > date_max:
                date_max = iso
            for c in series_cols:
                v = r[c] if c < len(r) else None
                if not isinstance(v, (int, float)):
                    continue
                fx = (str(fxs[c]).replace("FX=", "").strip() if c < len(fxs) and fxs[c] else "")
                key_tuple = (cat, str(names[c]).strip(), norm_ticker(tickers[c] if c < len(tickers) else None), fx)
                iid = lookup.get(key_tuple)
                if iid is None:
                    unmatched.add(key_tuple)
                    continue
                rows_buffer.append({"instrument_id": iid, "obs_date": iso, "value": float(v)})
                if len(rows_buffer) >= BATCH:
                    flush()

    flush()

    if unmatched:
        print(f"\nWARNING: {len(unmatched)} series columns had no matching instrument (skipped):")
        for k in list(unmatched)[:20]:
            print("   ", k)

    # Record the load in load_log.
    api("POST", url, key, "/rest/v1/load_log", body=[{
        "instruments_updated": len(lookup),
        "rows_written": total,
        "date_min": date_min,
        "date_max": date_max,
        "note": "Historical load from Data Dump - HP.xlsx",
    }], extra_headers={"Prefer": "return=minimal"})

    print(f"\nDONE. Loaded {total:,} datapoints  ({date_min} -> {date_max}).")


if __name__ == "__main__":
    main()
